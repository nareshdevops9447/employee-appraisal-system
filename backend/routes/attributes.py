"""
Attributes routes — management of behavioral competency templates.
This implements Tier 1 of the ownership model.
"""
from flask import Blueprint, request, jsonify, g, send_file
from werkzeug.utils import secure_filename
import openpyxl
from io import BytesIO
from extensions import db
from models.attribute_template import AttributeTemplate
from models.employee_attribute import EmployeeAttribute
from utils.decorators import require_auth, require_role

attributes_bp = Blueprint('attributes', __name__)

@attributes_bp.route('/cycle/<cycle_id>', methods=['GET'])
@require_auth
def list_cycle_attributes(cycle_id):
    """List all attribute templates for a specific cycle."""
    templates = AttributeTemplate.query.filter_by(cycle_id=cycle_id, is_active=True)\
        .order_by(AttributeTemplate.display_order.asc()).all()
    return jsonify([t.to_dict() for t in templates])

@attributes_bp.route('/', methods=['POST'])
@require_role('hr_admin', 'super_admin')
def create_attribute_template():
    """Create a new attribute template. HR Admin only."""
    data = request.get_json()
    if not data or not data.get('cycle_id') or not data.get('title'):
        return jsonify({'error': 'cycle_id and title are required'}), 400

    template = AttributeTemplate(
        cycle_id=data['cycle_id'],
        title=data['title'],
        description=data.get('description'),
        display_order=data.get('display_order', 0),
        created_by=g.current_user['user_id']
    )
    db.session.add(template)
    db.session.commit()
    return jsonify(template.to_dict()), 201

@attributes_bp.route('/<id>', methods=['PUT'])
@require_role('hr_admin', 'super_admin')
def update_attribute_template(id):
    """Update an attribute template. HR Admin only."""
    template = AttributeTemplate.query.get_or_404(id)
    data = request.get_json()

    if 'title' in data:
        template.title = data['title']
    if 'description' in data:
        template.description = data['description']
    if 'display_order' in data:
        template.display_order = data['display_order']
    if 'is_active' in data:
        template.is_active = data['is_active']

    db.session.commit()
    return jsonify(template.to_dict())

@attributes_bp.route('/<id>', methods=['DELETE'])
@require_role('hr_admin', 'super_admin')
def delete_attribute_template(id):
    """Soft delete (deactivate) an attribute template. HR Admin only."""
    template = AttributeTemplate.query.get_or_404(id)
    template.is_active = False
    db.session.commit()
    return jsonify({'message': 'Attribute template deactivated'})

@attributes_bp.route('/upload', methods=['POST'])
@require_role('hr_admin', 'super_admin')
def upload_attribute_templates():
    """Bulk upload attribute templates via an Excel file (.xlsx)"""
    cycle_id = request.form.get('cycle_id')
    if not cycle_id:
        return jsonify({'error': 'cycle_id is required'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'No file part in the request'}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
        
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Only .xlsx files are supported'}), 400

    try:
        # Load workbook from memory
        wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        sheet = wb.active

        # Check existing attributes count for this cycle
        existing_count = AttributeTemplate.query.filter_by(cycle_id=cycle_id, is_active=True).count()
        remaining_slots = 5 - existing_count

        if remaining_slots <= 0:
            return jsonify({'error': 'Cycle already has 5 active attributes. Delete some before uploading.'}), 400

        added_count = 0
        templates_to_add = []

        # Assuming Row 1 is header: Title, Description
        for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            if added_count >= remaining_slots:
                break # Stop if we reach the max limit of 5
                
            title = row[0]
            description = row[1] if len(row) > 1 else None

            if not title:
                continue # Skip empty rows

            template = AttributeTemplate(
                cycle_id=cycle_id,
                title=str(title).strip(),
                description=str(description).strip() if description else None,
                display_order=existing_count + added_count,
                created_by=g.current_user['user_id']
            )
            templates_to_add.append(template)
            added_count += 1
            
        if templates_to_add:
            db.session.add_all(templates_to_add)
            db.session.commit()
            
        return jsonify({
            'message': f'Successfully uploaded {added_count} attribute templates.',
            'count': added_count
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to process Excel file: {str(e)}'}), 500

@attributes_bp.route('/template/download', methods=['GET'])
def download_attribute_template():
    """Download a blank Excel template for attribute templates."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attributes"
    
    # Headers
    ws.append(["Title", "Description"])
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    
    # Save to BytesIO
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Attribute_Templates_Upload.xlsx'
    )

@attributes_bp.route('/employee-ratings/<employee_id>/<cycle_id>', methods=['GET'])
@require_auth
def get_employee_attribute_ratings(employee_id, cycle_id):
    """Get ratings for an employee's attributes in a cycle."""
    ctx = g.current_user
    # Only self, manager, or HR
    is_owner = ctx['user_id'] == employee_id
    is_hr = ctx['role'] in ('hr_admin', 'super_admin')
    
    # Check manager logic would go here if needed, keeping it simple for now
    if not (is_owner or is_hr):
        # Check manager status
        from models.user_profile import UserProfile
        profile = UserProfile.query.get(employee_id)
        if not (profile and profile.manager_id == ctx['user_id']):
             return jsonify({'error': 'Forbidden'}), 403

    ratings = EmployeeAttribute.query.filter_by(employee_id=employee_id, cycle_id=cycle_id).all()
    return jsonify([r.to_dict() for r in ratings])

@attributes_bp.route('/employee-ratings', methods=['POST'])
@require_auth
def rate_employee_attribute():
    """Submit a rating for an attribute."""
    data = request.get_json()
    ctx = g.current_user
    
    required = ['attribute_template_id', 'employee_id', 'cycle_id']
    if not all(k in data for k in required):
        return jsonify({'error': 'Missing required fields'}), 400

    # Find or create rating record
    rating = EmployeeAttribute.query.filter_by(
        attribute_template_id=data['attribute_template_id'],
        employee_id=data['employee_id'],
        cycle_id=data['cycle_id']
    ).first()

    if not rating:
        rating = EmployeeAttribute(
            attribute_template_id=data['attribute_template_id'],
            employee_id=data['employee_id'],
            cycle_id=data['cycle_id']
        )
        db.session.add(rating)

    # Permission check for rating
    is_employee = ctx['user_id'] == data['employee_id']
    
    # Logic: Employee can only set self_rating, Manager only manager_rating
    if is_employee:
        if 'self_rating' in data:
            rating.self_rating = data['self_rating']
        if 'self_comment' in data:
            rating.self_comment = data['self_comment']
    else:
        # Check if current user is the manager
        from models.user_profile import UserProfile
        profile = UserProfile.query.get(data['employee_id'])
        if profile and profile.manager_id == ctx['user_id']:
            if 'manager_rating' in data:
                rating.manager_rating = data['manager_rating']
            if 'manager_comment' in data:
                rating.manager_comment = data['manager_comment']
        else:
             return jsonify({'error': 'Only the assigned manager can rate this employee'}), 403

    db.session.commit()
    return jsonify(rating.to_dict())
