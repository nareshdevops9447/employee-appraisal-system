"""
Goal Templates routes — management of HR-defined performance goals.
This implements the mandatory performance goals for Phase 1-3.

Updated: department_id support so HR can create dept-specific templates.
Managers can fetch templates relevant to their team via /my-department/<cycle_id>.
"""
from flask import Blueprint, request, jsonify, g, send_file
from werkzeug.utils import secure_filename
import openpyxl
from io import BytesIO
from extensions import db
from models.goal_template import GoalTemplate
from models.user_profile import UserProfile
from utils.decorators import require_auth, require_role

goal_templates_bp = Blueprint('goal_templates', __name__)

@goal_templates_bp.route('/cycle/<cycle_id>', methods=['GET'])
@require_auth
def list_cycle_goal_templates(cycle_id):
    """List goal templates for a specific cycle.
    
    Optional query param: ?department_id=<id>
    - If provided, returns templates for that department + org-wide templates.
    - If not provided, returns all templates for the cycle.
    """
    dept_id = request.args.get('department_id')

    query = GoalTemplate.query.filter_by(cycle_id=cycle_id, is_active=True)

    if dept_id:
        # Return org-wide (NULL dept) + the specified department's templates
        query = query.filter(
            db.or_(
                GoalTemplate.department_id.is_(None),
                GoalTemplate.department_id == dept_id
            )
        )

    templates = query.order_by(GoalTemplate.display_order.asc()).all()
    return jsonify([t.to_dict() for t in templates])


@goal_templates_bp.route('/my-department/<cycle_id>', methods=['GET'])
@require_auth
def my_department_templates(cycle_id):
    """Returns templates for the current user's department + all org-wide templates.
    
    Designed for managers: they see both shared org-wide goals and
    department-specific ones relevant to their team.
    """
    ctx = g.current_user
    profile = UserProfile.query.get(ctx['user_id'])
    dept_id = profile.department_id if profile else None

    query = GoalTemplate.query.filter_by(cycle_id=cycle_id, is_active=True)

    if dept_id:
        query = query.filter(
            db.or_(
                GoalTemplate.department_id.is_(None),
                GoalTemplate.department_id == dept_id
            )
        )
    else:
        # No dept — only show org-wide
        query = query.filter(GoalTemplate.department_id.is_(None))

    templates = query.order_by(GoalTemplate.display_order.asc()).all()
    return jsonify([t.to_dict() for t in templates])


@goal_templates_bp.route('/', methods=['POST'])
@require_role('hr_admin', 'super_admin')
def create_goal_template():
    """Create a new goal template. HR Admin only."""
    data = request.get_json()
    if not data or not data.get('cycle_id') or not data.get('title'):
        return jsonify({'error': 'cycle_id and title are required'}), 400

    template = GoalTemplate(
        cycle_id=data['cycle_id'],
        title=data['title'],
        description=data.get('description'),
        category=data.get('category', 'performance'),
        display_order=data.get('display_order', 0),
        department_id=data.get('department_id'),  # None = org-wide
        created_by=g.current_user['user_id']
    )
    db.session.add(template)
    db.session.commit()
    return jsonify(template.to_dict()), 201

@goal_templates_bp.route('/<id>', methods=['PUT'])
@require_role('hr_admin', 'super_admin')
def update_goal_template(id):
    """Update a goal template. HR Admin only."""
    template = GoalTemplate.query.get_or_404(id)
    data = request.get_json()

    if 'title' in data:
        template.title = data['title']
    if 'description' in data:
        template.description = data['description']
    if 'category' in data:
        template.category = data['category']
    if 'display_order' in data:
        template.display_order = data['display_order']
    if 'is_active' in data:
        template.is_active = data['is_active']
    if 'department_id' in data:
        # Allow setting to None (org-wide) or a specific dept id
        template.department_id = data['department_id'] or None

    db.session.commit()
    return jsonify(template.to_dict())

@goal_templates_bp.route('/<id>', methods=['DELETE'])
@require_role('hr_admin', 'super_admin')
def delete_goal_template(id):
    """Soft delete (deactivate) a goal template. HR Admin only."""
    template = GoalTemplate.query.get_or_404(id)
    template.is_active = False
    db.session.commit()
    return jsonify({'message': 'Goal template deactivated'})

@goal_templates_bp.route('/upload', methods=['POST'])
@require_role('hr_admin', 'super_admin')
def upload_goal_templates():
    """Bulk upload goal templates via an Excel file (.xlsx)
    
    Excel columns (row 1 = header):
      A: Title          — required
      B: Description    — optional
      C: Category       — optional (defaults to 'performance')
      D: Department     — optional; leave blank for org-wide.
                          Must match an existing department name exactly.
    """
    from models.department import Department

    cycle_id = request.form.get('cycle_id')
    if not cycle_id:
        return jsonify({'error': 'cycle_id is required'}), 400

    if 'file' not in request.files:
        return jsonify({'error': 'No file part in the request'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'No selected file'}), 400
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Only .xlsx files are supported'}), 400

    try:
        wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
        sheet = wb.active

        # Pre-load all departments into a name→id map (case-insensitive)
        all_depts = Department.query.all()
        dept_map = {d.name.strip().lower(): d.id for d in all_depts}

        added_count = 0
        skipped_count = 0
        invalid_depts = []
        templates_to_add = []

        # Row 1 = header; columns: Title, Description, Category, Department
        for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
            title = row[0]
            description = row[1] if len(row) > 1 else None
            category = row[2] if len(row) > 2 and row[2] else 'performance'
            dept_name_raw = row[3] if len(row) > 3 else None

            if not title:
                continue

            title = str(title).strip()
            dept_name = str(dept_name_raw).strip() if dept_name_raw else None

            # Resolve department name → id
            department_id = None
            if dept_name:
                department_id = dept_map.get(dept_name.lower())
                if not department_id:
                    invalid_depts.append(dept_name)
                    continue  # Skip rows with unrecognised departments

            # Skip exact duplicates (same title + same dept scope)
            existing = GoalTemplate.query.filter_by(
                cycle_id=cycle_id,
                title=title,
                department_id=department_id
            ).first()
            if existing:
                skipped_count += 1
                continue

            template = GoalTemplate(
                cycle_id=cycle_id,
                title=title,
                description=str(description).strip() if description else None,
                category=str(category).strip().lower(),
                display_order=added_count,
                department_id=department_id,
                created_by=g.current_user['user_id']
            )
            templates_to_add.append(template)
            added_count += 1

        if templates_to_add:
            db.session.add_all(templates_to_add)
            db.session.commit()

        response = {
            'message': f'Successfully uploaded {added_count} goal template(s).',
            'created': added_count,
            'skipped_duplicates': skipped_count,
        }
        if invalid_depts:
            response['warnings'] = (
                f"{len(invalid_depts)} row(s) skipped — department name not found: "
                + ', '.join(set(invalid_depts))
                + ". Check the Departments sheet in the template for valid names."
            )
        return jsonify(response), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to process Excel file: {str(e)}'}), 500

@goal_templates_bp.route('/template/download', methods=['GET'])
def download_goal_template():
    """Download an Excel template for bulk goal template upload.
    
    Sheet 1 (Goals): Title, Description, Category, Department
    Sheet 2 (Departments): Reference list of all valid department names from the DB.
    """
    from models.department import Department
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Sheet 1: Goals ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Goals"

    # Style helpers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")  # indigo
    note_fill = PatternFill("solid", fgColor="EEF2FF")    # light indigo
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    headers = [
        "Title *",
        "Description",
        "Category",
        "Department (blank = Org-wide)",
    ]
    ws.append(headers)

    # Style header row
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Example rows
    examples = [
        ["Operational Excellence & Delivery",
         "Deliver assigned tasks to agreed standards and timelines.",
         "performance",
         ""],  # blank = org-wide
        ["Sales Pipeline Growth",
         "Achieve monthly sales target by growing the pipeline.",
         "performance",
         "Sales"],
        ["Engineering Sprint Delivery",
         "Complete sprint commitments with less than 10% carry-over.",
         "performance",
         "Engineering"],
        ["Team Learning & Development",
         "Complete at least one learning course per quarter.",
         "development",
         ""],
    ]
    for row_data in examples:
        ws.append(row_data)

    # Light background on example rows so they stand out from blank rows
    for row_idx in range(2, 2 + len(examples)):
        for col_idx in range(1, 5):
            ws.cell(row=row_idx, column=col_idx).fill = note_fill
            ws.cell(row=row_idx, column=col_idx).alignment = left_wrap

    # Column widths
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 30
    ws.row_dimensions[1].height = 22

    # Instruction row at the top (insert above header by shifting, easier to add note)
    ws.insert_rows(1)
    note_cell = ws.cell(row=1, column=1,
        value=(
            "⚠ Instructions: Fill Title (required). "
            "Leave Department blank for org-wide goals. "
            "Department names must exactly match those on the 'Departments' sheet."
        )
    )
    note_cell.font = Font(italic=True, color="374151")
    note_cell.fill = PatternFill("solid", fgColor="FEF9C3")  # yellow
    note_cell.alignment = left_wrap
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 36

    # ── Sheet 2: Departments reference ───────────────────────────────
    ws2 = wb.create_sheet("Departments")
    ws2.append(["Valid Department Names"])
    ws2.cell(row=1, column=1).font = Font(bold=True)
    ws2.column_dimensions['A'].width = 30

    try:
        depts = Department.query.order_by(Department.name.asc()).all()
        for dept in depts:
            ws2.append([dept.name])
    except Exception:
        ws2.append(["(Could not load departments — ensure DB is running)"])

    # ── Output ────────────────────────────────────────────────────────
    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Goal_Templates_Upload.xlsx'
    )
